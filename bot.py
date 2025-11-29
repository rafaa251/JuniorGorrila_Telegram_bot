from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from dotenv import load_dotenv
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")


PRODUСTS = [
    {"name": "3D фигурка", "price": 299},
    {"name": "Запчасть на автомобиль", "price": 4500},
    {"name": "3D игрушка", "price": 224},
    {"name": "Концелярия", "price": 259},
    {"name": "Школьные предметы", "price": 432},
]


def product_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🔙 Назад", callback_data="prev"),
            InlineKeyboardButton("➡️ Далее", callback_data="next"),
        ],
        [
            InlineKeyboardButton("🛒 В корзину", callback_data="in_basket")
        ],
        [
            InlineKeyboardButton("📝 Корзина", callback_data="view_cart")
        ]
    ])

def cart_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("💸 Оформить", callback_data="checkout"),
            InlineKeyboardButton("🗑️ Очистить корзину", callback_data="clear_cart"),
        ],
        [
            InlineKeyboardButton("🔙 Назад к товарам", callback_data="back_to_products")
        ]
    ])

def confirm_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Потвердить", callback_data="confirm")],
        [InlineKeyboardButton("❎ Отмена", callback_data="cancel")]
    ])


def cart_total_and_text(cart):
    if not cart:
        return 0, "🛒 Корзина пуста"
    
    lines = ["🛒 *Корзина:*"]
    total = 0
    
    for i, item in enumerate(cart, start=1):
        line_sum = item["price"] * item["qty"]
        lines.append(f"{i}) {item['name']} - {item['qty']} шт. - {line_sum} руб.")
        total += line_sum
        
    lines.append(f"\n*Итого:* {total} руб.")
    return total, "\n".join(lines)
    

async def show_product(update, context, index):
    product = PRODUСTS[index]
    text = (
        f"*🥡 {product['name']}*\n"
        f"💸 Цена: *{product['price']} руб.*\n\n"
        "Выберите действие ниже 👇🏿"
    )

    if update.callback_query:
        try:
            await update.callback_query.edit_message_text(
                text,
                parse_mode="Markdown",
                reply_markup=product_keyboard()
            )
        except Exception:
            await update.callback_query.message.reply_text(
                text,
                parse_mode="Markdown",
                reply_markup=product_keyboard()
            )
    else:
        await update.message.reply_text(
            text,
            parse_mode="Markdown",
            reply_markup=product_keyboard()
        )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Привет! Пиши /menu чтобы смотреть товары 🔍")

async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.setdefault("product_index", 0)
    context.user_data.setdefault("cart", [])
    context.user_data["stage"] = None
    await update.message.reply_text("Начинаем просмотр товаров...")
    await show_product(update, context, context.user_data["product_index"])

async def buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    index = context.user_data.get("product_index", 0)
    cart = context.user_data.setdefault("cart", [])

    if data == "next":
        index = (index + 1) % len(PRODUСTS)
        context.user_data["product_index"] = index
        await show_product(update, context, index)
        return

    if data == "prev":
        index = (index - 1) % len(PRODUСTS)
        context.user_data["product_index"] = index
        await show_product(update, context, index)
        return
    
    if data == "in_basket":
        context.user_data["state"] = "WAITING_QTY"
        context.user_data["pending_product_index"] = index
        try:
            await query.edit_message_text("Какое количество добавить в корзину? Введите целое число")
        except Exception:
            await query.edit_message_text("Какое количество вы хотите заказать? Укажите целое число!")
        return
    
    if data == "view_cart":
        total, text = cart_total_and_text(cart)
        if cart:
            inline_buttons = []
            for i, item in enumerate(cart):
                inline_buttons.append([InlineKeyboardButton(f"💣 Удалить: {item['name']}", callback_data=f"del:{i}")])
            inline_buttons.append([InlineKeyboardButton("💸 Оформить", callback_data="checkout"),
                                   InlineKeyboardButton("🗑️ Очистить корзину", callback_data="clear_cart")])
            inline_buttons.append([InlineKeyboardButton("🔙 Назад к товарам", callback_data="back_to_products")])
            kb = InlineKeyboardMarkup(inline_buttons)
            try:
                await query.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
            except Exception:
                await query.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)
        else:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад к товарам", callback_data="back_to_products")]])
            try:
                await query.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
            except Exception:
                await query.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)
        return
    
    if data.startswith("del:"):
        try:
            _, idx_str = data.split(":", 1)
            del_idx = int(idx_str)
            if 0 <= del_idx < len(cart):
                removed = cart.pop(del_idx)
                total, text = cart_total_and_text(cart)
                if cart:
                    inline_buttons = []
                    for i, item in enumerate(cart):
                        inline_buttons.append([InlineKeyboardButton(f"Удалить: {item['name']}", callback_data=f"del:{i}")])
                    inline_buttons.append([InlineKeyboardButton("Оформить", callback_data="checkout"),
                                            InlineKeyboardButton("Очистить корзину", callback_data="clear_cart")])
                    inline_buttons.append([InlineKeyboardButton("🔙 Назад к товарам", callback_data="back_to_products")])
                    kb = InlineKeyboardMarkup(inline_buttons)
                else:
                    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад к товарам", callback_data="back_to_products")]])
                try: 
                    await query.edit_message_text(f"Удалено: {removed['name']}\n\n{text}", parse_mode="Markdown", reply_markup=kb)
                except Exception:
                    await query.message.reply_text(f"Удалено: {removed['name']}\n\n{text}", parse_mode="Markdown", reply_markup=kb)
            else:
                await query.answer("Неверный индекс", show_alert=True)
        except Exception:
            await query.answer("Ошибка при удалении", show_alert=True)
        return
    if data == "clear_cart":
        context.user_data["cart"] = []
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад к товарам", callback_data="back_to_products")]])
        try:
            await query.message.edit_text("Корзина очищена", reply_markup=kb)
        except Exception:
            await query.message.edit_text("Корзина очищена", reply_markup=kb)
        return
    
    if data == "back_to_products":
        idx = context.user_data.get("product_index", 0)
        await show_product(update, context, idx)
        return
    
    if data == "checkout":
        if not cart:
            await query.answer("Корзина пустая", show_alert=True)
            return
        context.user_data["state"] = "WAITING_NAME"
        try:
            await query.edit_message_text("Введите ФИО для оормления заказа:")
        except Exception:
            await query.message.reply_text("Введите ФИО для оформления заказа:")
        return
    
    if data == "confirm":
        await create_excel_and_send(update, context)
        return
    
    if data == "cancel":
        context.user_data["state"] = None
        await query.edit_message_text("(Оформление отменено)")
        idx = context.user_data.get("product_index", 0)
        await show_product(update, context, idx)
        return
    
async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    state = context.user_data.get("state")
    if state == "WAITING_QTY":
        try:
            qty = int(update.message.text.strip())
            if qty <= 0:
                raise ValueError
        except Exception:
            await update.message.reply_text("Введите нормальное положительное число")
            return
        
        pending_idx = context.user_data.get("pending_product_index")
        if pending_idx is None or not (0 <= pending_idx < len(PRODUСTS)):
            await update.message.reply_text("Не удалось определить товар, попробуй снова")
            context.user_data["state"] = None
            context.user_data.pop("pending_product_index", None)
            return
        
        product = PRODUСTS[pending_idx]
        cart = context.user_data.setdefault("cart", [])
        found = False
        for it in cart:
            if it['name'] == product['name']:
                it["qty"] += qty
                found = True
                break
        if not found:
            cart.append({"name": product["name"], "price": product["price"], "qty": qty})

        context.user_data["state"] = None
        context.user_data.pop("pending_product_index", None)
        await update.message.reply_text(f"Добавлено: {product['name']} — {qty} шт.")
        await show_product(update, context, context.user_data.get("product_index", 0))
        return
    
    if state == "WAITING_NAME":
        context.user_data["customer_name"] = update.message.text.strip()
        context.user_data["state"] = "CONFIRM"
        total, text = cart_total_and_text(context.user_data.get("cart", []))
        confirm_text = f"Проверьте данные:\n\nФИО: {context.user_data['customer_name']}\n\n{text}\n\nНажмите потвердить чтобы завершить"
        await update.message.reply_text(confirm_text, parse_mode="Markdown", reply_markup=confirm_keyboard())
        return
    await update.message.reply_text("Не понял тебя, используй кнопки или для просмотра товаров нажми /menu")

async def create_excel_and_send(update: Update, context):
    query = update.callback_query if update.callback_query else None
    name = context.user_data.get("customer_name", "Не указан")
    cart = context.user_data.get("cart", [])

    filename = "order.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Дата", "ФИО", "Товар", "Количество"])
        wb.save(filename)

    wb = load_workbook(filename)
    sheet = wb.active
    for item in cart:
        sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            name,
            item["name"],
            item["qty"]
        ])
    wb.save(filename)

    try:
        if query and query.message:
            await query.message.reply_document(open(filename, "rb"), caption="Ваш заказ оформлен")
        else:
            await update.effective_chat.send_document(open(filename, "rb"), caption="Ваш заказ оформлен")
    except Exception as e:
        try:
            await update.effective_chat.send_document(open(filename, "rb"), caption="Ваш заказ оформлен")
        except Exception:
            pass

    context.user_data.clear()


if __name__ == "__main__":
    app = ApplicationBuilder().token(BOT_TOKEN).build()
   
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", menu))
    app.add_handler(CallbackQueryHandler(buttons))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    print("Бот начал работу.......")
    app.run_polling()

    
    

    

    



